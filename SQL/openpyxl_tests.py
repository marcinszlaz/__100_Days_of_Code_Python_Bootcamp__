import openpyxl

sheet_as_a_dict = {}

# opening workbook (with open no needed :])
workbook = openpyxl.load_workbook(filename="testy.xlsx",read_only = True,keep_vba = True)
sheet = workbook["Arkusz1"]

# calculate dimensions xD
print('dimensions of sheet',sheet.calculate_dimension())

# creating header of table in dictionary
def make_header():
  """makes header (dict) from first row in sheet"""
  global sheet_as_a_dict
  h_l=[]
  smart_dict = {}
  for row_one in range(1,2):
    for col in range(1,sheet.max_column+1):
      head = sheet.cell(row=row_one,column=col)
      h_l.append(head.value)
  smart_dict = {f'col_{i}': h for i,h in enumerate(h_l,1)}
  sheet_as_a_dict["header"] = smart_dict
  return f"{sheet_as_a_dict}"

# filling rows
def make_body():
  """ function assumes first row is a header
      and make body from other rows(s)  """
  tmp_list = []
  tmp_dict = {}
  for row in range(2,sheet.max_row+1):
    for col in range(1,sheet.max_column+1):
      cell_content = sheet.cell(row=row,column=col).value
      tmp_list.append(cell_content)
    tmp_dict = {f"col_{i}": cell for i,cell in enumerate(tmp_list,1)}
    sheet_as_a_dict[f"row_{row}"]= tmp_dict
  return f"{sheet_as_a_dict}"

print('header',make_header())
print('header + body',make_body())